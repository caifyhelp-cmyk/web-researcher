# -*- coding: utf-8 -*-
"""
범용 웹 리서치 어시스턴트 v4
──────────────────────────────────
- 네이버 + 구글 동시 검색
- GPT 리서치 플랜 수립 (A/B/C 타입 + 동적 제외목록 + 필터강도)
- A타입: 특정 도메인 내부 다중 페이지 크롤링
- B/C타입: 멀티엔진 수집 → 배치 필터링 → 자동 재시도 (결과 부족 시)
- XLSX 저장
"""

import os, time, re, json, random
try:
    from duckduckgo_search import DDGS
    _DDG_OK = True
except ImportError:
    _DDG_OK = False
import requests
from bs4 import BeautifulSoup
from urllib.parse import quote, urlparse, urljoin, unquote
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from openai import OpenAI

# ──────────────────────────────────────────────
#  설정
# ──────────────────────────────────────────────
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
NAVER_DELAY = 2.0
GOOGLE_DELAY = 2.5
SELENIUM_DELAY = 3.0
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

NAVER_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "ko-KR,ko;q=0.9",
    "Referer": "https://www.naver.com",
}
GOOGLE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# 절대 제외 (포털·SNS·위키 — 어떤 조사에도 의미 없음)
HARD_EXCLUDE = [
    "naver.com", "navercorp.com", "google.com", "daum.net", "bing.com", "kakao.com",
    "youtube.com", "youtu.be",
    "instagram.com", "facebook.com", "twitter.com", "x.com", "linkedin.com",
    "namu.wiki", "wikipedia.org",
]

# 블로그·콘텐츠 플랫폼 — strict_filter=high/medium 일 때 추가 차단
BLOG_PLATFORM_EXCLUDE = [
    "velog.io", "brunch.co.kr", "tistory.com", "blog.naver.com",
    "wikidocs.net", "publy.co", "mobiinside.co.kr",
    "ppss.kr", "sharedit.co.kr", "earticle.net", "dbpia.co.kr",
    "repository.hanyang.ac.kr", "scienceon.kisti.re.kr",
]


# ══════════════════════════════════════════════
#  채팅 인터페이스
# ══════════════════════════════════════════════

def gpt_suggest_needs(client, keyword):
    prompt = f"""
사용자가 웹에서 다음을 조사하려 합니다: "{keyword}"

조사 목적을 판단해서 실무에 바로 쓸 수 있는 수집 항목 5~7개 추천:
- 경쟁사/랜딩페이지 분석 → 마케팅·UX·전환 관점
- 특정 기관/서비스 분석 → 서비스 구조·프로세스·가격
- 연락처/데이터 수집 → 실무 데이터 항목
- 뉴스/트렌드 → 핵심 내용·시사점
각 항목 10자 이내. JSON 배열로만: ["항목1", ...]
"""
    try:
        r = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.5, max_tokens=200,
        )
        m = re.search(r'\[.*?\]', r.choices[0].message.content, re.DOTALL)
        if m:
            return json.loads(m.group())
    except Exception:
        pass
    return ["주요 내용 요약", "특징", "가격", "차별점", "연락처"]


def gpt_suggest_context(client, keyword, needs):
    try:
        r = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": f"""
조사 주제: "{keyword}"  /  수집 항목: {', '.join(needs)}
이 조사를 더 정확하게 하기 위한 조건이나 방향을 1~2문장으로 제안. 제안문만 답하세요.
"""}],
            temperature=0.5, max_tokens=150,
        )
        return r.choices[0].message.content.strip()
    except Exception:
        return ""


def chat_interface(client):
    print("\n" + "=" * 64)
    print("   범용 웹 리서치 어시스턴트 v4")
    print("=" * 64)

    print("\n─" * 64)
    print("Q1. 무엇을 조사할까요?")
    print("    예) 관리감독자 교육 경쟁사 / 대한산업안전협회 수강신청")
    print("    예) 전국 대학 안전팀 연락처 / AI 자동화 툴 비교")
    keyword = input("\n    → ").strip()
    if not keyword:
        return chat_interface(client)

    print("\n  [분석 중...]")
    suggested_needs = gpt_suggest_needs(client, keyword)

    print("\n─" * 64)
    print("Q2. 어떤 정보가 필요한가요?")
    print(f"\n  추천 항목:")
    for i, s in enumerate(suggested_needs, 1):
        print(f"    {i}) {s}")
    print("\n  Enter=전체  /  번호(예:1,3,5)  /  직접입력")
    needs_raw = input("\n    → ").strip()

    if not needs_raw:
        needs = suggested_needs
    elif re.match(r'^[\d,\s]+$', needs_raw):
        idx = [int(x.strip())-1 for x in needs_raw.split(",") if x.strip().isdigit()]
        needs = [suggested_needs[i] for i in idx if 0 <= i < len(suggested_needs)] or suggested_needs
    else:
        needs = [n.strip() for n in needs_raw.replace("，",",").split(",") if n.strip()]

    print("\n  [분석 중...]")
    suggested_context = gpt_suggest_context(client, keyword, needs)

    print("\n─" * 64)
    print("Q3. 추가 조건이나 방향이 있나요?")
    if suggested_context:
        print(f"\n  제안: {suggested_context}")
        print("\n  Enter=제안사용  /  직접입력  /  skip=없음")
    ctx_raw = input("\n    → ").strip()
    context = ("" if ctx_raw.lower()=="skip"
               else ctx_raw if ctx_raw
               else suggested_context)

    print("\n─" * 64)
    print("Q4. 몇 개 수집할까요? (기본 10, 최대 25)")
    count_str = input("\n    → ").strip()
    try:
        count = max(3, min(int(count_str), 25))
    except Exception:
        count = 10

    print("\n" + "━" * 64)
    print(f"  조사 주제 : {keyword}")
    print(f"  수집 항목 : {' / '.join(needs)}")
    if context:
        print(f"  추가 조건 : {context}")
    print(f"  수집 목표 : {count}개")
    print("━" * 64)
    if input("\n  시작? (Enter=예 / n=다시): ").strip().lower() == "n":
        return chat_interface(client)

    return {"keyword": keyword, "count": count, "needs": needs, "context": context}


# ══════════════════════════════════════════════
#  리서치 플랜 수립
# ══════════════════════════════════════════════

def build_research_plan(client, config, model="gpt-4o-mini"):
    """
    GPT가 조사 유형(A/B/C) + 검색 전략 + 동적 제외목록 + 필터강도 수립.
    쿼리를 네이버용 4개 + 구글용 4개 따로 생성 (각 엔진 특성 반영).
    model: 기본 gpt-4o-mini, 고품질 시 gpt-4o 전달
    """
    prompt = f"""
조사: "{config['keyword']}"
항목: {', '.join(config['needs'])}
맥락: {config.get('context','없음')}

━━ 조사 유형 판단 ━━
A) 특정 기관·사이트 집중 분석 (예: "대한산업안전협회 수강신청", "삼성SDS 서비스")
B) 시장·경쟁사 비교 (예: "관리감독자 교육 경쟁사", "AI 툴 비교")
C) 특정 정보·데이터 수집 (예: "대학 안전팀 연락처", "최신 뉴스")

━━ 추가 판단 기준 ━━
- UI/UX·사용성·기능비교 조사면 strict_filter는 반드시 "medium" (블로그·리뷰 포함 필요)
- 뉴스·기사 수집 조사면 strict_filter="low", exclude_extra=[]
- 연락처·목록 등 대량 기관 데이터 수집이면 is_bulk=true, entity_queries에 기관명 열거 쿼리 추가

━━ 쿼리 생성 규칙 ━━
- 쿼리는 실제 고객이 네이버/구글에 입력할 법한 자연스러운 검색어
- "분석", "전환율", "랜딩 페이지 최적화" 같은 마케터 용어 절대 금지
- 목적이 기관·서비스 탐색이면 "[키워드] 기관", "[키워드] 협회", "[키워드] 신청", "[키워드] 업체 추천" 형태로
- 목적이 경쟁사 탐색이면 "[키워드] 비교", "[키워드] 추천", "[키워드] 업체" 형태로
- 구글 쿼리는 영문 또는 한영 혼합 가능

━━ 아래 JSON만 출력 ━━
{{
  "type": "A or B or C",
  "reason": "판단 이유 한 줄",
  "target_domain": "A타입 대상 도메인 (예: kosa.or.kr), 아니면 빈 문자열",
  "naver_queries": ["네이버용 쿼리1", "쿼리2", "쿼리3", "쿼리4"],
  "google_queries": ["구글용 쿼리1(영문가능)", "쿼리2", "쿼리3", "쿼리4"],
  "entity_queries": ["기관명 열거 쿼리 (대량 수집 시만, 예: '서울대 안전팀 연락처', '연세대 안전팀')"],
  "exclude_extra": ["추가 제외 도메인 패턴 (뉴스 조사면 반드시 빈 배열)"],
  "strict_filter": "high(서비스사이트만) or medium(관련사이트 폭넓게) or low(뉴스·정보사이트 포함)",
  "is_bulk": false
}}
"""
    try:
        r = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3, max_tokens=500,
        )
        raw = r.choices[0].message.content.strip()
        m = re.search(r'\{[\s\S]+\}', raw)
        if m:
            plan = json.loads(m.group())
            print(f"  유형    : {plan['type']}타입 — {plan.get('reason','')}")
            if plan.get("target_domain"):
                print(f"  대상    : {plan['target_domain']}")
            print(f"  네이버  : {plan.get('naver_queries', [])}")
            print(f"  구글    : {plan.get('google_queries', [])}")
            print(f"  필터    : {plan.get('strict_filter','medium')}")
            return plan
    except Exception as e:
        print(f"  [!] 플랜 오류: {e}")

    kw = config['keyword']
    return {
        "type": "B", "reason": "기본값", "target_domain": "",
        "naver_queries": [kw, f"{kw} 서비스", f"{kw} 추천", f"{kw} 비교"],
        "google_queries": [kw, f"{kw} site comparison", f"best {kw}", f"{kw} review"],
        "exclude_extra": [], "strict_filter": "medium",
    }


# ══════════════════════════════════════════════
#  멀티엔진 검색
# ══════════════════════════════════════════════

def search_naver(query, max_pages=2):
    results, seen = [], set()
    for page in range(1, max_pages + 1):
        try:
            resp = requests.get(
                f"https://search.naver.com/search.naver?query={quote(query)}&where=web&start={(page-1)*10+1}",
                headers=NAVER_HEADERS, timeout=8
            )
            soup = BeautifulSoup(resp.content.decode("utf-8", errors="replace"), "html.parser")
        except Exception:
            break

        blocks = (soup.select(".total_wrap") or soup.select(".lst_total li")
                  or soup.select("[data-cr-id]") or [soup])
        for block in blocks:
            for a in block.select("a[href]"):
                href = a.get("href", "")
                if "url=" in href and "naver" in href:
                    m = re.search(r'url=([^&]+)', href)
                    if m: href = unquote(m.group(1))
                if not href.startswith("http"): continue
                domain = urlparse(href).netloc.lower()
                if domain in seen: continue
                title = a.get_text(strip=True) or domain
                if len(title) < 2: continue
                seen.add(domain)
                results.append({"url": href, "title": title, "domain": domain, "source": "naver"})
        time.sleep(NAVER_DELAY)
    return results


def search_google(query, max_results=15):
    results, seen = [], set()
    try:
        resp = requests.get(
            f"https://www.google.com/search?q={quote(query)}&num={max_results}&hl=ko&gl=kr",
            headers=GOOGLE_HEADERS, timeout=10
        )
        soup = BeautifulSoup(resp.content.decode("utf-8", errors="replace"), "html.parser")

        # 여러 selector 시도 (Google 구조 변경 대응)
        for sel in ["div.g", "div.tF2Cxc", "div.MjjYud > div", "div[data-sokoban-container]"]:
            blocks = soup.select(sel)
            if not blocks:
                continue
            for block in blocks:
                a_tag = block.select_one("a[href]")
                h3    = block.select_one("h3")
                if not a_tag or not h3:
                    continue
                href = a_tag.get("href", "")
                if href.startswith("/url?q="):
                    href = unquote(href[7:].split("&")[0])
                if not href.startswith("http"):
                    continue
                domain = urlparse(href).netloc.lower()
                if domain in seen:
                    continue
                seen.add(domain)
                results.append({
                    "url": href,
                    "title": h3.get_text(strip=True),
                    "domain": domain,
                    "source": "google"
                })
            if results:
                break

        # fallback: 모든 외부 링크
        if not results:
            for a_tag in soup.select("a[href]"):
                href = a_tag.get("href", "")
                if href.startswith("/url?q="):
                    href = unquote(href[7:].split("&")[0])
                if not href.startswith("http") or "google" in href:
                    continue
                domain = urlparse(href).netloc.lower()
                if domain in seen or not domain:
                    continue
                seen.add(domain)
                results.append({
                    "url": href,
                    "title": a_tag.get_text(strip=True) or domain,
                    "domain": domain,
                    "source": "google"
                })
        time.sleep(GOOGLE_DELAY)
    except Exception as e:
        print(f"    [!] Google 오류: {e}")
    return results[:max_results]




def search_duckduckgo(query, max_results=15):
    """DuckDuckGo 검색 — 서버 IP 차단 없음, 무료"""
    if not _DDG_OK:
        return []
    results, seen = [], set()
    try:
        with DDGS() as ddgs:
            for r in ddgs.text(query, region='kr-kr', max_results=max_results):
                url   = r.get('href', '')
                title = r.get('title', '') or r.get('body', '')[:60]
                if not url.startswith('http'):
                    continue
                domain = urlparse(url).netloc.lower()
                if domain in seen:
                    continue
                seen.add(domain)
                results.append({"url": url, "title": title, "domain": domain, "source": "ddg"})
    except Exception as e:
        print(f"    [!] DDG 오류: {e}")
    return results

def collect_candidates(plan, exclude_all, collect_target):
    """네이버 + 구글로 후보 URL 수집 (중복 제거, is_bulk 시 entity_queries 추가)"""
    urls_seen  = set()
    candidates = []

    naver_queries  = plan.get("naver_queries", [])
    google_queries = plan.get("google_queries", [])
    entity_queries = plan.get("entity_queries", []) if plan.get("is_bulk") else []

    # entity_queries를 naver_queries 앞에 추가 (대량 수집 시 기관명 직접 검색 우선)
    if entity_queries:
        print("  [대량수집] entity_queries 사용")
        naver_queries = entity_queries + naver_queries

    print("  [네이버 검색]")
    for q in naver_queries:
        if len(candidates) >= collect_target: break
        print(f"    → '{q}'")
        for item in search_naver(q):
            if len(candidates) >= collect_target: break
            domain = item["domain"]
            bare   = domain.replace("www.", "")
            if any(bare == ex or bare.endswith("."+ex) for ex in exclude_all): continue
            if domain in urls_seen: continue
            urls_seen.add(domain)
            candidates.append(item)
            print(f"      [후보] {domain}")

    print("  [구글 검색]")
    for q in google_queries:
        if len(candidates) >= collect_target: break
        print(f"    → '{q}'")
        for item in search_google(q):
            if len(candidates) >= collect_target: break
            domain = item["domain"]
            bare   = domain.replace("www.", "")
            if any(bare == ex or bare.endswith("."+ex) for ex in exclude_all): continue
            if domain in urls_seen: continue
            urls_seen.add(domain)
            candidates.append(item)
            print(f"      [후보] {domain} (Google)")

    # ── DuckDuckGo 폴백: 네이버+구글 결과 부족 시 ──────────────────
    if len(candidates) < collect_target // 2 and _DDG_OK:
        print("  [DDG 폴백] 네이버/구글 결과 부족 → DuckDuckGo 검색")
        all_queries = (naver_queries + google_queries)[:4]
        for q in all_queries:
            if len(candidates) >= collect_target: break
            print(f"    → DDG '{q}'")
            for item in search_duckduckgo(q):
                if len(candidates) >= collect_target: break
                domain = item["domain"]
                bare   = domain.replace("www.", "")
                if any(bare == ex or bare.endswith("."+ex) for ex in exclude_all): continue
                if domain in urls_seen: continue
                urls_seen.add(domain)
                candidates.append(item)
                print(f"      [DDG후보] {domain}")

    return candidates


def filter_relevant_sites(client, candidates, config, plan):
    if not candidates:
        return []

    strictness_map = {
        "high":   (
            "실제로 이 서비스를 직접 제공하는 기관·기업·협회·단체 사이트만 선택. "
            "블로그·개인블로그·뉴스·리뷰·학술논문·플랫폼·해외 SaaS·식품·부동산 등 "
            "주제와 무관한 모든 것 제외."
        ),
        "medium": (
            "조사 목적과 직접 관련된 사이트. "
            "개인 블로그(tistory, velog, brunch), 무관 업종, 해외 SaaS는 제외. "
            "공식 기관·업체·플랫폼·리뷰 사이트 포함."
        ),
        "low":    "조사 목적과 조금이라도 관련 있으면 포함. 검색엔진·SNS만 제외.",
    }
    strictness = strictness_map.get(plan.get("strict_filter","medium"), strictness_map["medium"])

    lines = "\n".join(f"{i+1}. {c['domain']} | {c['title']}" for i,c in enumerate(candidates))
    prompt = f"""
조사 목적: "{config['keyword']}"
수집 항목: {', '.join(config['needs'])}
선택 기준: {strictness}

후보 목록:
{lines}

위 기준에 맞는 번호만 JSON 배열로 출력. 조금이라도 애매하면 제외.
예시: [1, 3, 5]
"""
    try:
        r = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0, max_tokens=200,
        )
        raw = r.choices[0].message.content.strip()
        m = re.search(r'\[.*?\]', raw, re.DOTALL)
        if m:
            indices = json.loads(m.group())
            return [candidates[i-1] for i in indices if 1 <= i <= len(candidates)]
    except Exception as e:
        print(f"  [!] 필터링 오류: {e}")
    return candidates


def search_all_engines(config, plan, client, existing_domains=None):
    """멀티엔진 수집 → 필터링 → 자동 재시도"""
    target_count   = config["count"]
    collect_target = min(target_count * 4, 80)
    strict = plan.get("strict_filter", "medium")
    blog_block = BLOG_PLATFORM_EXCLUDE if strict in ("high", "medium") else []
    exclude_all    = HARD_EXCLUDE + blog_block + plan.get("exclude_extra", [])
    existing_domains = existing_domains or set()

    candidates = collect_candidates(plan, exclude_all, collect_target)

    print(f"\n  후보 {len(candidates)}개 → GPT 필터링 중...")
    filtered = filter_relevant_sites(client, candidates, config, plan)
    # 이미 수집된 도메인 제거
    filtered = [f for f in filtered if f["domain"] not in existing_domains]
    filtered = filtered[:target_count]

    # ── 자동 재시도: 목표의 70% 미달이면 추가 쿼리 생성 ──
    retry = 0
    while len(filtered) < int(target_count * 0.7) and retry < 2:
        retry += 1
        print(f"\n  [재시도 {retry}] 결과 부족 ({len(filtered)}개) → 추가 쿼리 생성 중...")
        collected_domains = {f["domain"] for f in filtered}

        extra_prompt = f"""
조사 목적: "{config['keyword']}"
기존 수집 도메인: {list(collected_domains)[:10]}
아직 {target_count - len(filtered)}개가 더 필요합니다.

규칙:
- 실제 고객이 네이버/구글에 검색할 자연스러운 검색어
- "분석", "전환율", "최적화", "랜딩 페이지" 같은 마케터 용어 금지
- 기관·서비스 탐색이면 "[키워드] 기관", "[키워드] 업체", "[키워드] 추천" 형태
- 기존 도메인과 겹치지 않는 새로운 각도의 쿼리

JSON: {{"naver": ["q1","q2","q3"], "google": ["q1","q2","q3"]}}
"""
        try:
            r = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": extra_prompt}],
                temperature=0.6, max_tokens=200,
            )
            raw = r.choices[0].message.content.strip()
            m = re.search(r'\{[\s\S]+\}', raw)
            if m:
                extra = json.loads(m.group())
                extra_plan = {
                    **plan,
                    "naver_queries": extra.get("naver", []),
                    "google_queries": extra.get("google", []),
                }
                existing = {f["domain"] for f in filtered}
                extra_candidates = collect_candidates(extra_plan, exclude_all, collect_target // 2)
                extra_filtered   = filter_relevant_sites(client, extra_candidates, config, plan)
                extra_filtered   = [f for f in extra_filtered if f["domain"] not in existing]
                filtered.extend(extra_filtered)
                filtered = filtered[:target_count]
        except Exception as e:
            print(f"  [!] 재시도 오류: {e}")
            break

    print(f"\n  확정 {len(filtered)}개")
    for i, r in enumerate(filtered, 1):
        print(f"    [{i:2d}] {r['domain']} ({r.get('source','')})")
    return filtered


# ══════════════════════════════════════════════
#  A타입: 특정 도메인 내부 탐색
# ══════════════════════════════════════════════

def discover_internal_pages(driver, base_url, config, client, max_pages=10):
    domain = urlparse(base_url).netloc
    print(f"\n  [A타입] {domain} 내부 페이지 탐색...")

    visited, candidates = set(), []
    try:
        driver.get(base_url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(SELENIUM_DELAY)

        soup = BeautifulSoup(driver.page_source, "html.parser")
        for a in soup.find_all("a", href=True):
            href = urljoin(base_url, a["href"])
            if urlparse(href).netloc != domain or href in visited:
                continue
            label = a.get_text(strip=True)
            candidates.append({"url": href, "label": label})
            visited.add(href)
    except Exception as e:
        print(f"  [!] 탐색 오류: {e}")

    if not candidates:
        return [{"url": base_url, "title": "홈페이지", "domain": domain}]

    lines = "\n".join(f"{i+1}. {c['url']} | {c['label']}" for i,c in enumerate(candidates[:60]))
    prompt = f"""
조사: "{config['keyword']}"  /  항목: {', '.join(config['needs'])}
대상 사이트: {domain}

아래 내부 링크 중 조사에 가장 유용한 페이지 최대 {max_pages}개 번호 선택.
(서비스소개, 요금, 신청절차, 조직/부서, 연락처 등 목적에 맞는 것)

{lines}

번호만 JSON: [1, 5, 12]
"""
    selected = []
    try:
        r = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0, max_tokens=150,
        )
        raw = r.choices[0].message.content.strip()
        m = re.search(r'\[.*?\]', raw, re.DOTALL)
        if m:
            indices = json.loads(m.group())
            selected = [
                {"url": candidates[i-1]["url"], "title": candidates[i-1]["label"], "domain": domain}
                for i in indices if 1 <= i <= len(candidates)
            ]
    except Exception as e:
        print(f"  [!] 페이지 선별 오류: {e}")

    pages = [{"url": base_url, "title": "홈페이지", "domain": domain}]
    for p in selected:
        if p["url"] != base_url:
            pages.append(p)

    print(f"  → {len(pages)}개 페이지 선별")
    for p in pages:
        print(f"    - {p['title']} : {p['url'][:60]}")
    return pages[:max_pages]


# ══════════════════════════════════════════════
#  Selenium
# ══════════════════════════════════════════════

def _scrape_requests(url: str) -> dict:
    """Chrome 없는 서버 환경용 requests + BS4 폴백 스크래퍼"""
    try:
        resp = requests.get(url, headers=NAVER_HEADERS, timeout=10)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
            tag.decompose()
        title = soup.title.string.strip() if soup.title else ""
        text  = " ".join(soup.get_text(" ", strip=True).split())[:4000]
        return {"url": url, "page_title": title, "full_text": text,
                "meta_desc": "", "h1": "", "error": False}
    except Exception as e:
        return {"url": url, "page_title": "", "full_text": "",
                "error": True, "error_msg": str(e)}


_CHROME_OK = None

def make_driver():
    """Chrome 사용 가능하면 Selenium driver 반환, 없으면 None (requests 모드)"""
    global _CHROME_OK
    if _CHROME_OK is False:
        return None
    import platform, shutil
    try:
        opts = Options()
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--window-size=1920,1080")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        if platform.system() == "Linux":
            chrome = shutil.which("chromium") or shutil.which("chromium-browser") or shutil.which("google-chrome")
            cdriver = shutil.which("chromedriver")
            if not chrome or not cdriver:
                print("[서버] Chrome 없음 — requests 모드로 실행")
                _CHROME_OK = False
                return None
            opts.binary_location = chrome
            service = Service(cdriver)
        else:
            opts.add_experimental_option("excludeSwitches", ["enable-automation"])
            opts.add_experimental_option("useAutomationExtension", False)
            service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=opts)
        driver.execute_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        _CHROME_OK = True
        return driver
    except Exception as e:
        print(f"[Chrome 오류] {e} — requests 모드 전환")
        _CHROME_OK = False
        return None


def validate_domain(url):
    """HEAD 요청으로 도메인 접근 가능 여부 확인"""
    try:
        resp = requests.head(url, timeout=6, allow_redirects=True, headers=NAVER_HEADERS)
        return resp.status_code < 400
    except Exception:
        return False


def is_login_wall(page_data):
    """페이지가 로그인 장벽인지 감지"""
    text = (page_data.get("full_text", "") + " " + page_data.get("page_title", "")).lower()
    signals = [
        "로그인이 필요", "로그인 후", "로그인하세요", "로그인 후 이용",
        "회원가입이 필요", "회원만 이용",
        "sign in to", "please log in", "login required", "login to continue",
        "you must be logged in",
    ]
    return any(s in text for s in signals)


def scrape_page(driver, url):
    if driver is None:
        return _scrape_requests(url)
    empty = {"page_title": "", "headings": "", "full_text": "", "error": ""}
    try:
        driver.get(url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(SELENIUM_DELAY)

        soup = BeautifulSoup(driver.page_source, "html.parser")
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()

        headings = [
            f"[{t.name.upper()}] {t.get_text(strip=True)}"
            for t in soup.find_all(["h1","h2","h3"])[:25]
            if t.get_text(strip=True)
        ]
        full_text = re.sub(r"\s+", " ", soup.get_text(" ")).strip()[:3000]
        return {
            "page_title": (driver.title or "")[:120],
            "headings":   "\n".join(headings),
            "full_text":  full_text,
            "error":      "",
        }
    except Exception as e:
        empty["error"] = str(e)[:120]
        return empty


# ══════════════════════════════════════════════
#  GPT 분석
# ══════════════════════════════════════════════

def analyze_with_gpt(client, url, page_data, config):
    needs     = config["needs"]
    needs_str = "\n".join(f'  "{n}": ""' for n in needs)
    prompt = f"""
웹 리서치 전문가로서 아래 페이지를 분석해주세요.

[조사 주제]: {config['keyword']}
[추가 맥락]: {config.get('context') or '없음'}
[URL]: {url}
[제목]: {page_data['page_title']}
[헤딩]:
{page_data['headings']}
[본문]:
{page_data['full_text']}

없으면 "확인 불가". 각 항목 2~5문장.
JSON만:
{{
  "한줄요약": "",
{needs_str}
}}
"""
    try:
        r = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3, max_tokens=1500,
        )
        raw = r.choices[0].message.content.strip()
        m = re.search(r"\{[\s\S]+\}", raw)
        if m:
            return json.loads(m.group())
    except Exception as e:
        print(f"    [!] GPT 오류: {e}")
    return {"한줄요약": "분석 실패", **{n: "" for n in needs}}


# ══════════════════════════════════════════════
#  XLSX 저장
# ══════════════════════════════════════════════

def save_xlsx(all_results, config, research_type="B"):
    needs    = config["needs"]
    keyword  = config["keyword"]
    THIN     = Side(style="thin", color="CCCCCC")

    def cell(ws, row, col, value, bold=False, bg=None, fc="000000", wrap=True, size=10, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=bold, color=fc, size=size, name="맑은 고딕")
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
        c.border    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        if bg: c.fill = PatternFill("solid", fgColor=bg)
        return c

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "리서치 결과"

    fixed = ([("No",5),("페이지명",20),("URL",32),("페이지 제목",26),("한줄 요약",30)]
             if research_type == "A"
             else [("No",5),("출처",8),("URL",32),("도메인",18),("페이지 제목",26),("한줄 요약",30)])

    all_headers = fixed + [(n, 30) for n in needs] + [("헤딩 구조",28),("오류",14)]
    total_cols  = len(all_headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    tc = ws.cell(row=1, column=1,
                 value=f"웹 리서치 [{research_type}타입] — {keyword}  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    tc.font      = Font(bold=True, size=13, color="FFFFFF", name="맑은 고딕")
    tc.fill      = PatternFill("solid", fgColor="1F3864")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, (hdr, width) in enumerate(all_headers, 1):
        cell(ws, 2, ci, hdr, bold=True, bg="2F5496", fc="FFFFFF", align="center")
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 22

    for ri, data in enumerate(all_results, 1):
        er  = ri + 2
        bg  = "EBF0FA" if ri % 2 else "FFFFFF"
        gpt = data.get("gpt", {})
        pg  = data.get("page", {})
        ws.row_dimensions[er].height = 80

        if research_type == "A":
            row_vals = [ri, data.get("title",""), data.get("url",""),
                        pg.get("page_title",""), gpt.get("한줄요약","")]
        else:
            row_vals = [ri, data.get("source",""), data.get("url",""),
                        data.get("domain",""), pg.get("page_title",""), gpt.get("한줄요약","")]

        for n in needs:
            row_vals.append(gpt.get(n, ""))
        row_vals += [pg.get("headings",""), pg.get("error","")]

        for ci, val in enumerate(row_vals, 1):
            cell(ws, er, ci, str(val) if val else "", bg=bg,
                 align="center" if ci == 1 else "left")

    ws.freeze_panes = "B3"
    safe_kw  = re.sub(r'[\\/:*?"<>|]', "_", keyword)
    filepath = os.path.join(OUTPUT_DIR, f"리서치_{safe_kw}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(filepath)
    print(f"\n  [저장] {filepath}")
    return filepath


# ══════════════════════════════════════════════
#  메인
# ══════════════════════════════════════════════

def main():
    client = OpenAI(api_key=OPENAI_API_KEY)
    config = chat_interface(client)

    print("\n[플랜] 리서치 계획 수립 중...")
    plan = build_research_plan(client, config)
    research_type = plan.get("type", "B")

    driver, all_results = None, []

    try:
        # ── A타입: 특정 사이트 내부 깊게 ──
        if research_type == "A":
            target_domain = plan.get("target_domain", "")
            if target_domain:
                base_url = (target_domain if target_domain.startswith("http")
                            else f"https://www.{target_domain}")
            else:
                print("\n  [A타입] 대상 도메인 검색 중...")
                temp = search_all_engines(config, {**plan, "count": 1}, client)
                base_url = temp[0]["url"] if temp else ""

            if not base_url:
                print("  [!] 대상 사이트 미발견 → B타입으로 전환")
                research_type = "B"
            else:
                # ── 사전 검증: 도메인 접근 가능 여부 확인 ──
                print(f"\n  [검증] {base_url} 접근 확인 중...")
                if not validate_domain(base_url):
                    print("  [!] 접근 불가 (404/타임아웃) → B타입으로 전환")
                    research_type = "B"
                else:
                    print(f"\n[준비] Chrome 드라이버 초기화...")
                    driver = make_driver()

                    # 홈 페이지 선 렌더링해서 로그인 장벽 체크
                    print("  [검증] 로그인 장벽 확인 중...")
                    home_data = scrape_page(driver, base_url)
                    if is_login_wall(home_data):
                        print("  [!] 로그인 필요 사이트 감지 → 공개 정보 검색으로 전환 (B타입)")
                        research_type = "B"
                    else:
                        pages = discover_internal_pages(driver, base_url, config, client, config["count"])

                        for i, pg in enumerate(pages, 1):
                            print(f"\n[{i:2d}/{len(pages)}] {pg['url'][:65]}")
                            print("  → 렌더링...")
                            page_data = scrape_page(driver, pg["url"])
                            if page_data["error"]:
                                print(f"  [!] {page_data['error'][:60]}")
                                continue
                            if is_login_wall(page_data):
                                print("  [스킵] 로그인 장벽 페이지")
                                continue
                            print("  → GPT 분석...")
                            gpt_result = analyze_with_gpt(client, pg["url"], page_data, config)
                            print(f"  [OK] {gpt_result.get('한줄요약','')[:55]}")
                            all_results.append({**pg, "page": page_data, "gpt": gpt_result})
                            time.sleep(random.uniform(1.0, 2.0))

        # ── B/C타입: 멀티엔진 수집 ──
        if research_type in ("B", "C"):
            pages = search_all_engines(config, plan, client)
            if not pages:
                print("[오류] 수집된 URL 없음")
                return

            if driver is None:
                print(f"\n[준비] Chrome 드라이버 초기화...")
                driver = make_driver()
            else:
                print(f"\n[준비] 기존 드라이버 재사용")

            for i, pg in enumerate(pages, 1):
                print(f"\n[{i:2d}/{len(pages)}] {pg['domain']} ({pg.get('source','')})")
                print("  → 렌더링...")
                page_data = scrape_page(driver, pg["url"])
                if page_data["error"]:
                    print(f"  [!] {page_data['error'][:60]}")
                    try: driver.quit(); driver = make_driver()
                    except Exception: pass
                    continue
                if is_login_wall(page_data):
                    print("  [스킵] 로그인 장벽 — 공개 정보 없음")
                    continue
                print("  → GPT 분석...")
                gpt_result = analyze_with_gpt(client, pg["url"], page_data, config)
                print(f"  [OK] {gpt_result.get('한줄요약','')[:55]}")
                all_results.append({**pg, "page": page_data, "gpt": gpt_result})
                time.sleep(random.uniform(1.5, 2.5))

    finally:
        if driver:
            try: driver.quit()
            except Exception: pass

    if not all_results:
        print("[오류] 분석 결과 없음")
        return

    filepath = save_xlsx(all_results, config, research_type)
    print("\n" + "=" * 64)
    print(f"  완료!  타입:{research_type}  /  {len(all_results)}개 분석")
    print(f"  결과: {filepath}")
    print("=" * 64)


if __name__ == "__main__":
    main()
